"""
잡플래닛 커뮤니티 크롤러 + 감정 분석 통합
URL  : https://www.jobplanet.co.kr/community/home
출력 : jobplanet_community_YYYYMMDD_HHMMSS.xlsx  (단일 파일)

동작 순서
  [1/4] 잡플래닛 접속 (세션 초기화)
  [2/4] API 호출로 게시글 수집
  [3/4] GPT-4o 감정 분석 (긍정 / 부정 / 중립 + 분류 이유)
  [4/4] Excel 저장
"""

import os
import re
import json
import time
from datetime import datetime
from urllib.parse import urlencode

import pandas as pd
from dotenv import load_dotenv
from openai import OpenAI
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ── 크롤러 설정 ───────────────────────────────────────────────────────────────
CATEGORY_ID: int | None = None   # None=전체 / 1=이직·취준 / 2=아무얘기 / 3=직장생활 / 4=출퇴근톡
MAX_POSTS: int = 100              # 수집할 최대 글 수 (0 = 무제한)
PAGE_LIMIT: int = 20              # 한 번에 가져올 글 수 (최대 20)
ORDER_BY: str = "recent"          # "recent" | "popular"
HEADLESS: bool = True             # False 로 바꾸면 브라우저가 보임
CRAWL_PAUSE: float = 0.5          # 크롤링 API 호출 간격 (초)

# ── 감정 분석 설정 ────────────────────────────────────────────────────────────
SENTIMENT_PAUSE: float = 1.0      # GPT API 호출 간격 (초)
MAX_CONTENT_LEN: int = 600        # 분석할 본문 최대 글자 수
MAX_RETRIES: int = 5              # Rate Limit 재시도 횟수
# ─────────────────────────────────────────────────────────────────────────────

BASE_URL = "https://www.jobplanet.co.kr"
API_URL  = f"{BASE_URL}/api/v5/community/posts"

CATEGORY_MAP = {
    None: "전체",
    1: "이직/취준",
    2: "아무얘기",
    3: "직장생활",
    4: "출퇴근톡",
}

SENTIMENT_COLORS = {
    "긍정":   ("DFF2BF", "4CAF50"),
    "부정":   ("FFD2D2", "C0392B"),
    "중립":   ("FFF9C4", "827717"),
    "분류불가": ("E0E0E0", "616161"),
}

SYSTEM_PROMPT = """당신은 커뮤니티 게시글 감정 분석 전문가입니다.
주어진 게시글을 아래 기준으로 분류하고, 반드시 JSON 형식으로만 출력하세요.

분류 기준:
- 긍정: 칭찬이라 판단되는 내용이 포함된 글
- 부정: 불평이나 불만의 내용이 포함된 글
- 중립: 긍정과 부정이 모두 포함되거나 어느 쪽에도 해당하지 않는 글

출력 형식 (반드시 아래 JSON만 출력, 다른 텍스트 없이):
{"감정": "긍정/부정/중립 중 하나", "이유": "분류 근거를 1~2문장으로 설명"}"""


# ── OpenAI 클라이언트 초기화 ──────────────────────────────────────────────────
load_dotenv()
_openai_client: OpenAI | None = None

def get_openai_client() -> OpenAI:
    global _openai_client
    if _openai_client is None:
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise RuntimeError(".env 파일에 OPENAI_API_KEY 가 없습니다.")
        _openai_client = OpenAI(api_key=api_key)
    return _openai_client


# ── 크롤링 ────────────────────────────────────────────────────────────────────

def build_params(cursor: str | None = None) -> dict:
    p: dict = {"limit": PAGE_LIMIT, "order_by": ORDER_BY}
    if CATEGORY_ID is not None:
        p["category_id"] = CATEGORY_ID
    if cursor:
        p["cursor"] = cursor
    return p


def flatten_post(item: dict) -> dict:
    cat          = item.get("community_category") or {}
    company_tags = item.get("company_tags") or []
    images       = item.get("images") or []
    author_str   = item.get("author_employee_status_snapshot_str") or ""
    employee_str = item.get("employee_status_str") or ""

    post_id    = item.get("id", "")
    share_link = item.get("share_link", "")
    post_url   = share_link if share_link else f"{BASE_URL}/community/posts/{post_id}"

    content         = (item.get("content") or "").strip()
    content_preview = content[:200].replace("\n", " ")

    return {
        "ID":         post_id,
        "카테고리":   cat.get("name", ""),
        "내용":       content,
        "내용 미리보기": content_preview,
        "작성자 정보": author_str or employee_str,
        "작성 시각":  item.get("created_at_str", ""),
        "좋아요":     item.get("likes_count", 0),
        "댓글수":     item.get("comments_count", 0),
        "조회수":     item.get("views_count", 0),
        "이미지 수":  len(images),
        "태그 기업":  ", ".join(c.get("name", "") for c in company_tags if isinstance(c, dict)),
        "게시글 URL": post_url,
    }


def fetch_all_posts(page_obj) -> list[dict]:
    all_posts: list[dict] = []
    cursor: str | None = None
    page_num = 0

    while True:
        page_num += 1
        params = build_params(cursor)
        url = f"{API_URL}?{urlencode(params)}"
        print(f"  페이지 {page_num:3d} 호출: ...{url[-60:]}", end="  ")

        try:
            resp = page_obj.request.get(url, headers={"Accept": "application/json"})
            if resp.status != 200:
                print(f"→ HTTP {resp.status}, 중단")
                break
            body = resp.json()
        except Exception as e:
            print(f"→ 오류: {e}, 중단")
            break

        data  = body.get("data") or {}
        items = data.get("items") or []
        cursor = data.get("cursor")

        if not items:
            print("→ 빈 응답, 종료")
            break

        for item in items:
            all_posts.append(flatten_post(item))

        collected = len(all_posts)
        print(f"→ {len(items)}개 수신 (누적: {collected}개)")

        if MAX_POSTS and collected >= MAX_POSTS:
            print(f"  목표 {MAX_POSTS}개 달성, 중단")
            break

        if not cursor or len(items) < PAGE_LIMIT:
            print("  마지막 페이지 도달")
            break

        time.sleep(CRAWL_PAUSE)

    return all_posts[:MAX_POSTS] if MAX_POSTS else all_posts


# ── 감정 분석 ─────────────────────────────────────────────────────────────────

def classify(content: str) -> tuple[str, str]:
    """GPT-4o 로 감정 분류 + 이유 반환.
    Returns: (감정 레이블, 분류 이유)
    """
    client = get_openai_client()
    text = content.strip()[:MAX_CONTENT_LEN]
    if not text:
        return "중립", "내용 없음"

    for attempt in range(MAX_RETRIES):
        try:
            resp = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user",   "content": text},
                ],
                max_tokens=150,
                temperature=0,
                response_format={"type": "json_object"},
            )
            raw    = resp.choices[0].message.content.strip()
            parsed = json.loads(raw)
            label  = parsed.get("감정", "").strip()
            reason = parsed.get("이유", "").strip()

            if label not in ("긍정", "부정", "중립"):
                for candidate in ("긍정", "부정", "중립"):
                    if candidate in label:
                        label = candidate
                        break
                else:
                    label = "중립"

            return label, reason

        except Exception as e:
            err = str(e)
            if "429" in err or "rate_limit" in err.lower():
                wait = 2 ** attempt
                m = re.search(r"try again in (\d+)ms", err)
                if m:
                    wait = max(int(m.group(1)) / 1000 + 0.5, wait)
                print(f"\n  [Rate Limit] {wait:.1f}초 후 재시도 ({attempt+1}/{MAX_RETRIES})...")
                time.sleep(wait)
            else:
                print(f"\n  [API 오류] {e}")
                return "분류불가", ""

    return "분류불가", ""


def run_sentiment(df: pd.DataFrame) -> pd.DataFrame:
    """DataFrame 의 각 행을 GPT-4o 로 분류, 감정·분류 이유 컬럼 추가 후 반환."""
    total = len(df)
    sentiments, reasons = [], []
    counts = {"긍정": 0, "부정": 0, "중립": 0, "분류불가": 0}

    for _, row in df.iterrows():
        content = str(row.get("내용", row.get("내용 미리보기", "")))
        label, reason = classify(content)
        sentiments.append(label)
        reasons.append(reason)
        counts[label] = counts.get(label, 0) + 1

        done   = len(sentiments)
        filled = int(30 * done / total)
        bar    = "#" * filled + "-" * (30 - filled)
        print(
            f"\r  [{bar}] {done:4d}/{total}  "
            f"긍정:{counts['긍정']} 부정:{counts['부정']} 중립:{counts['중립']}",
            end="", flush=True,
        )
        time.sleep(SENTIMENT_PAUSE)

    print()

    df = df.copy()
    df["감정"]    = sentiments
    df["분류 이유"] = reasons

    # 카테고리 바로 뒤에 감정·분류 이유 배치
    cols = list(df.columns)
    for col in ("분류 이유", "감정"):
        if col in cols:
            cols.remove(col)
    if "카테고리" in cols:
        idx = cols.index("카테고리") + 1
        cols.insert(idx, "감정")
        cols.insert(idx + 1, "분류 이유")
    df = df[cols]

    print(f"  긍정: {counts['긍정']}개  부정: {counts['부정']}개  "
          f"중립: {counts['중립']}개  분류불가: {counts.get('분류불가', 0)}개")

    return df, counts


# ── Excel 저장 ────────────────────────────────────────────────────────────────

COL_WIDTHS = {
    "ID": 10, "카테고리": 14, "감정": 10, "분류 이유": 55,
    "내용 미리보기": 50, "내용": 60,
    "작성자 정보": 20, "작성 시각": 18,
    "좋아요": 10, "댓글수": 10, "조회수": 10, "이미지 수": 10,
    "태그 기업": 24, "게시글 URL": 55,
}

COL_ORDER = [
    "ID", "카테고리", "감정", "분류 이유",
    "내용 미리보기", "내용",
    "작성자 정보", "작성 시각",
    "좋아요", "댓글수", "조회수", "이미지 수",
    "태그 기업", "게시글 URL",
]


def save_excel(df: pd.DataFrame, filename: str):
    existing = [c for c in COL_ORDER if c in df.columns]
    extra    = [c for c in df.columns if c not in COL_ORDER]
    df = df[existing + extra]

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="커뮤니티")
        ws = writer.sheets["커뮤니티"]

        # 헤더
        hdr_fill = PatternFill("solid", fgColor="4472C4")
        hdr_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # 열 너비
        for i, col in enumerate(df.columns, start=1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = COL_WIDTHS.get(col, 18)

        # 데이터 행 기본
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=False, vertical="center")

        # 분류 이유 열: 줄바꿈
        if "분류 이유" in df.columns:
            col_idx = df.columns.get_loc("분류 이유") + 1
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        # 행 높이
        has_reason = "분류 이유" in df.columns
        for i in range(2, len(df) + 2):
            ws.row_dimensions[i].height = 40 if has_reason else 18

    # 감정 컬럼 색상
    if "감정" in df.columns:
        wb = load_workbook(filename)
        ws = wb.active
        sent_col = next(
            (i for i, cell in enumerate(ws[1], 1) if cell.value == "감정"), None
        )
        if sent_col:
            for row_idx in range(2, len(df) + 2):
                cell = ws.cell(row=row_idx, column=sent_col)
                bg, fg = SENTIMENT_COLORS.get(cell.value or "", ("FFFFFF", "000000"))
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(color=fg, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        wb.save(filename)


# ── 메인 ──────────────────────────────────────────────────────────────────────

def main():
    cat_label = CATEGORY_MAP.get(CATEGORY_ID, f"ID={CATEGORY_ID}")
    print("=" * 60)
    print("  잡플래닛 커뮤니티 크롤러 + 감정 분석")
    print(f"  카테고리: {cat_label}  |  최대 글 수: {MAX_POSTS or '무제한'}")
    print(f"  정렬: {ORDER_BY}")
    print("=" * 60)

    # ── [1/4] 크롤링 ──────────────────────────────────────────────────────────
    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=HEADLESS,
            args=["--lang=ko-KR", "--no-sandbox"],
        )
        ctx = browser.new_context(
            locale="ko-KR",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        page = ctx.new_page()

        print("\n[1/4] 잡플래닛 접속 (세션 초기화)...")
        page.goto(f"{BASE_URL}/community/home", wait_until="domcontentloaded", timeout=30_000)
        time.sleep(2)

        print("[2/4] API 호출로 게시글 수집 중...")
        posts = fetch_all_posts(page)
        browser.close()

    if not posts:
        print("\n[오류] 수집된 게시글이 없습니다.")
        return

    df = pd.DataFrame(posts)
    print(f"  수집 완료: {len(df)}개\n")

    # ── [3/4] 감정 분석 ───────────────────────────────────────────────────────
    print("[3/4] GPT-4o 감정 분석 중...")
    try:
        df, counts = run_sentiment(df)
    except RuntimeError as e:
        print(f"  [경고] 감정 분석 건너뜀: {e}")
        counts = {}

    # ── [4/4] 저장 ────────────────────────────────────────────────────────────
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"jobplanet_community_{ts}.xlsx"

    print(f"\n[4/4] Excel 저장 중: {filename}")
    save_excel(df, filename)

    total = len(df)
    print(f"\n[완료]")
    print(f"  저장 파일  : {filename}")
    print(f"  수집 글 수 : {total}개  |  카테고리: {cat_label}")
    if counts:
        print(f"  긍정: {counts.get('긍정',0)}개  "
              f"부정: {counts.get('부정',0)}개  "
              f"중립: {counts.get('중립',0)}개")


if __name__ == "__main__":
    main()
