"""
잡플래닛 커뮤니티 감정 분석기
- 가장 최신 jobplanet_community_*.xlsx 파일을 읽어
- GPT-4o 로 각 게시글을 긍정 / 부정 / 중립 분류
- 결과를 새 컬럼으로 추가한 Excel 파일로 저장
"""

import os
import glob
import time
from pathlib import Path
from dotenv import load_dotenv
from openai import OpenAI
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ── 설정 ─────────────────────────────────────────────────────────────────────
BATCH_PAUSE = 1.0       # API 호출 간 대기 (초) – rate limit 방지
MAX_CONTENT_LEN = 600   # 분석에 사용할 본문 최대 글자 수 (토큰 절약)
TEST_LIMIT = 0         # 테스트용 제한 (0 = 전체 처리)
# ─────────────────────────────────────────────────────────────────────────────

load_dotenv()
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

SYSTEM_PROMPT = """당신은 커뮤니티 게시글 감정 분석 전문가입니다.
주어진 게시글을 아래 기준으로 분류하고, 반드시 JSON 형식으로만 출력하세요.

분류 기준:
- 긍정: 칭찬이라 판단되는 내용이 포함된 글
- 부정: 불평이나 불만의 내용이 포함된 글
- 중립: 긍정과 부정이 모두 포함되거나 어느 쪽에도 해당하지 않는 글

출력 형식 (반드시 아래 JSON만 출력, 다른 텍스트 없이):
{"감정": "긍정/부정/중립 중 하나", "이유": "분류 근거를 1~2문장으로 설명"}"""


def classify(content: str, max_retries: int = 5) -> tuple[str, str]:
    """GPT-4o 로 감정 분류 + 이유 반환. Rate Limit 시 지수 백오프로 재시도.
    Returns: (감정 레이블, 분류 이유)
    """
    import json
    import re as _re

    text = content.strip()[:MAX_CONTENT_LEN]
    if not text:
        return "중립", "내용 없음"

    for attempt in range(max_retries):
        try:
            resp = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user",   "content": text},
                ],
                max_tokens=150,
                temperature=0,
                response_format={"type": "json_object"},
            )
            raw = resp.choices[0].message.content.strip()
            parsed = json.loads(raw)
            label = parsed.get("감정", "").strip()
            reason = parsed.get("이유", "").strip()

            if label not in ("긍정", "부정", "중립"):
                for candidate in ("긍정", "부정", "중립"):
                    if candidate in label:
                        label = candidate
                        break
                else:
                    label = "중립"

            return label, reason

        except Exception as e:
            err = str(e)
            if "429" in err or "rate_limit" in err.lower():
                wait = 2 ** attempt
                m = _re.search(r"try again in (\d+)ms", err)
                if m:
                    wait = max(int(m.group(1)) / 1000 + 0.5, wait)
                print(f"\n  [Rate Limit] {wait:.1f}초 후 재시도 ({attempt+1}/{max_retries})...")
                time.sleep(wait)
            else:
                print(f"\n  [API 오류] {e}")
                return "분류불가", ""

    return "분류불가", ""


def find_latest_excel() -> Path | None:
    # 감정분석 결과 파일은 제외하고 원본 크롤링 파일만 검색
    files = sorted(
        [f for f in glob.glob("jobplanet_community_*.xlsx") if "감정분석" not in f],
        key=os.path.getmtime,
        reverse=True,
    )
    return Path(files[0]) if files else None


def apply_sentiment_style(filepath: str, df: pd.DataFrame):
    """감정 컬럼에 색상 강조 적용."""
    wb = load_workbook(filepath)
    ws = wb.active

    colors = {
        "긍정": ("DFF2BF", "4CAF50"),  # (배경, 글자)
        "부정": ("FFD2D2", "C0392B"),
        "중립": ("FFF9C4", "827717"),
        "분류불가": ("E0E0E0", "616161"),
    }

    # 감정 열 인덱스 찾기
    sentiment_col = None
    for i, cell in enumerate(ws[1], start=1):
        if cell.value == "감정":
            sentiment_col = i
            break

    if sentiment_col is None:
        wb.save(filepath)
        return

    for row_idx in range(2, len(df) + 2):
        cell = ws.cell(row=row_idx, column=sentiment_col)
        label = cell.value or ""
        bg, fg = colors.get(label, ("FFFFFF", "000000"))
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(color=fg, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(filepath)


def main():
    print("=" * 60)
    print("  잡플래닛 커뮤니티 감정 분석 (GPT-4o)")
    print("=" * 60)

    # 최신 엑셀 파일 탐색
    src = find_latest_excel()
    if src is None:
        print("[오류] jobplanet_community_*.xlsx 파일이 없습니다.")
        print("       crawler.py 를 먼저 실행해 주세요.")
        return

    print(f"\n[1/4] 입력 파일: {src.name}")
    df = pd.read_excel(src, sheet_name="커뮤니티")
    if TEST_LIMIT:
        df = df.head(TEST_LIMIT).copy()
        print(f"      [테스트 모드] 상위 {TEST_LIMIT}개만 처리")
    total = len(df)
    print(f"      게시글 수 : {total}개")

    # 이미 분석된 파일이면 덮어쓰기 확인
    if "감정" in df.columns:
        print("      '감정' 컬럼이 이미 존재합니다. 재분석합니다.")

    print(f"\n[2/4] GPT-4o 감정 분류 시작...")
    sentiments = []
    reasons = []
    counts = {"긍정": 0, "부정": 0, "중립": 0, "분류불가": 0}

    for i, row in df.iterrows():
        content = str(row.get("내용", row.get("내용 미리보기", "")))
        label, reason = classify(content)
        sentiments.append(label)
        reasons.append(reason)
        counts[label] = counts.get(label, 0) + 1

        bar_len = 30
        filled = int(bar_len * len(sentiments) / total)
        bar = "#" * filled + "-" * (bar_len - filled)
        print(
            f"\r  [{bar}] {len(sentiments):4d}/{total}  "
            f"긍정:{counts['긍정']} 부정:{counts['부정']} 중립:{counts['중립']}",
            end="",
            flush=True,
        )
        time.sleep(BATCH_PAUSE)

    print()  # 줄바꿈

    # DataFrame 에 감정·이유 컬럼 삽입 (카테고리 바로 다음)
    df["감정"] = sentiments
    df["분류 이유"] = reasons
    cols = list(df.columns)
    if "카테고리" in cols:
        for col in ("분류 이유", "감정"):
            if col in cols:
                cols.remove(col)
        idx = cols.index("카테고리") + 1
        cols.insert(idx, "감정")
        cols.insert(idx + 1, "분류 이유")
        df = df[cols]

    # 저장
    out_name = src.stem + "_감정분석.xlsx"
    print(f"\n[3/4] Excel 저장 중: {out_name}")

    col_widths = {
        "ID": 10, "카테고리": 14, "감정": 10, "분류 이유": 55,
        "내용 미리보기": 50, "내용": 60,
        "작성자 정보": 20, "작성 시각": 18,
        "좋아요": 10, "댓글수": 10, "조회수": 10, "이미지 수": 10,
        "태그 기업": 24, "게시글 URL": 55,
    }

    with pd.ExcelWriter(out_name, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="커뮤니티")
        ws = writer.sheets["커뮤니티"]

        # 헤더 스타일
        hdr_fill = PatternFill("solid", fgColor="4472C4")
        hdr_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # 열 너비
        for i, col in enumerate(df.columns, start=1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = col_widths.get(col, 18)

        # 데이터 행 기본 스타일
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=False, vertical="center")

        # 분류 이유 열: 줄바꿈 허용
        reason_col_letter = None
        for i, col in enumerate(df.columns, start=1):
            if col == "분류 이유":
                reason_col_letter = ws.cell(1, i).column_letter
                break
        if reason_col_letter:
            for row in ws.iter_rows(min_row=2,
                                    min_col=df.columns.get_loc("분류 이유") + 1,
                                    max_col=df.columns.get_loc("분류 이유") + 1):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        for i in range(2, len(df) + 2):
            ws.row_dimensions[i].height = 40

    # 감정 컬럼 색상 강조
    apply_sentiment_style(out_name, df)

    print(f"\n[4/4] 완료!")
    print(f"   저장 파일  : {out_name}")
    print(f"   총 게시글  : {total}개")
    print(f"   긍정       : {counts['긍정']}개  ({counts['긍정']/total*100:.1f}%)")
    print(f"   부정       : {counts['부정']}개  ({counts['부정']/total*100:.1f}%)")
    print(f"   중립       : {counts['중립']}개  ({counts['중립']/total*100:.1f}%)")
    if counts.get("분류불가"):
        print(f"   분류불가   : {counts['분류불가']}개")


if __name__ == "__main__":
    main()
